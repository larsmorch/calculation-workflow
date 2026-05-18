import openpyxl
import re
import sys
import os

def get_defined_names_mapping(wb):
    """Creates a mapping from cell coordinates to defined names and input/output labels."""
    mapping = {}
    defined_names = getattr(wb, 'defined_names', None)
    if defined_names is None:
        return mapping

    for dn in defined_names.values():
        if not hasattr(dn, 'destinations'):
            continue

        raw_comment = getattr(dn, 'comment', None)
        if not isinstance(raw_comment, str):
            continue

        label = raw_comment.strip().lower()
        if label not in ('input', 'output'):
            continue

        try:
            destinations = list(dn.destinations)
        except TypeError:
            continue

        for destination in destinations:
            if not destination or len(destination) != 2:
                continue

            sheet_name, coord = destination
            if not sheet_name or not coord:
                continue

            # Remove absolute references ($) for easier matching
            clean_coord = coord.replace('$', '')
            mapping[f"{sheet_name}!{clean_coord}"] = {
                'name': dn.name,
                'label': label,
            }
    return mapping


def format_value_with_excel_decimal(value, number_format):
    """Format numeric output values using the cell's Excel number format."""
    if number_format is None or not isinstance(value, (int, float)):
        return str(value)

    format_code = str(number_format).split(';', 1)[0]
    format_code = re.sub(r'\[[^\]]*\]', '', format_code)
    format_code = re.sub(r'"[^"]*"', '', format_code)
    format_code = format_code.strip()
    if not format_code or format_code.lower() == 'general':
        return str(value)

    is_percent = '%' in format_code
    if is_percent:
        value *= 100
        format_code = format_code.replace('%', '')

    decimal_match = re.search(r'\.([0#]+)', format_code)
    decimals = len(decimal_match.group(1)) if decimal_match else 0
    use_comma = ',' in format_code.split('.')[0]

    if decimals > 0:
        formatted = f"{value:{',' if use_comma else ''}.{decimals}f}"
    else:
        formatted = f"{value:{',' if use_comma else ''}.0f}"

    if is_percent:
        formatted += '%'
    return formatted

def extract_variables(excel_file_path):
    if not os.path.exists(excel_file_path):
        print(f"Error: File '{excel_file_path}' does not exist.")
        return [], []

    print(f"Reading {excel_file_path}...")
    try:
        # Load workbook, data_only=True reads the values instead of formulas
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    except Exception as e:
        print(f"Error loading {excel_file_path}: {e}")
        return [], []

    name_mapping = get_defined_names_mapping(wb)
    
    inputs = []
    outputs = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        for row in sheet.iter_rows():
            for cell in row:
                cell_ref = f"{sheet_name}!{cell.coordinate}"
                defined_info = name_mapping.get(cell_ref)

                comment_label = None
                if cell.comment is not None:
                    raw_text = getattr(cell.comment, 'text', '')
                    if isinstance(raw_text, str):
                        comment_label = raw_text.strip().lower()

                if comment_label in ('input', 'output'):
                    label = comment_label
                    var_name = defined_info['name'] if defined_info else cell_ref
                elif defined_info is not None:
                    label = defined_info['label']
                    var_name = defined_info['name']
                else:
                    continue

                unit_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                unit_value = unit_cell.value if unit_cell is not None else None
                unit = str(unit_value).strip() if isinstance(unit_value, str) and unit_value.strip() else None

                item = {
                    'name': var_name,
                    'type': label,
                    'value': cell.value,
                    'unit': unit,
                    'number_format': getattr(cell, 'number_format', None),
                }

                if label == 'input':
                    inputs.append(item)
                elif label == 'output':
                    outputs.append(item)

    return inputs, outputs

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_variables.py <path_to_excel_file.xlsx>")
        print("\nNote: You may need to install openpyxl if you haven't already:")
        print("pip install openpyxl")
    else:
        inputs, outputs = extract_variables(sys.argv[1])
        print("\n" + "="*30)
        print(f"--- INPUTS ({len(inputs)}) ---")
        if not inputs:
            print("No inputs found.")
        for item in inputs:
            formatted_value = format_value_with_excel_decimal(item['value'], item.get('number_format'))
            unit_text = f" {item['unit']}" if item.get('unit') else ''
            print(f"{item['name']} = {formatted_value}{unit_text}")

        print("\n" + "="*30)
        print(f"--- OUTPUTS ({len(outputs)}) ---")
        if not outputs:
            print("No outputs found.")
        for item in outputs:
            formatted_value = format_value_with_excel_decimal(item['value'], item.get('number_format'))
            unit_text = f" {item['unit']}" if item.get('unit') else ''
            print(f"{item['name']} = {formatted_value}{unit_text}")
